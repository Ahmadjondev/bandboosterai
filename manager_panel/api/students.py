"""
Manager API - Student Management Endpoints
"""

import re
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.shortcuts import get_object_or_404
from django.contrib.auth import get_user_model
from django.db.models import Q
from django.http import HttpResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

from ielts.models import ExamAttempt
from ..serializers import UserSerializer, UserDetailSerializer
from .utils import (
    check_manager_permission,
    permission_denied_response,
    paginate_queryset,
)

# Import score calculation helper
from .exams import _calculate_attempt_scores

User = get_user_model()


def normalize_phone(phone):
    """
    Normalize phone number by removing all symbols and ensuring + prefix
    Examples:
        '998911234567' -> '+998911234567'
        '+998911234567' -> '+998911234567'
        '998.91.123.45.67' -> '+998911234567'
        '998-91-123-45-67' -> '+998911234567'
        '998 91 123 45 67' -> '+998911234567'
        '+998 (91) 123-45-67' -> '+998911234567'
        '998,91,123,45,67' -> '+998911234567'
    """
    if not phone:
        return phone

    # Remove all non-digit and non-plus characters
    cleaned = re.sub(r"[^\d+]", "", phone)

    # Remove all + symbols first
    cleaned = cleaned.replace("+", "")

    # Add + at the beginning if there are digits
    if cleaned:
        return "+" + cleaned
    print(phone)
    return phone


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_students_list(request):
    """Get list of students"""
    if not check_manager_permission(request.user):
        return permission_denied_response()

    search = request.GET.get("search", "")

    students = User.objects.filter(role="STUDENT")

    if search:
        students = students.filter(
            Q(username__icontains=search)
            | Q(email__icontains=search)
            | Q(first_name__icontains=search)
            | Q(last_name__icontains=search)
        )

    students = students.order_by("-date_joined")

    paginated = paginate_queryset(students, request)
    serializer = UserSerializer(paginated["results"], many=True)
    return Response(
        {"students": serializer.data, "pagination": paginated["pagination"]}
    )


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_student_detail(request, user_id):
    """Get detailed student information"""
    if not check_manager_permission(request.user):
        return permission_denied_response()

    student = get_object_or_404(User, id=user_id, role="STUDENT")

    # Get completed exam attempts for this student
    attempts = (
        ExamAttempt.objects.filter(student=student, status="COMPLETED")
        .select_related("exam", "exam__mock_test")
        .order_by("-completed_at")
    )

    # Build results with calculated scores
    results_data = []
    for attempt in attempts:
        scores = _calculate_attempt_scores(attempt)

        results_data.append(
            {
                "id": attempt.id,
                "exam_title": attempt.exam.name,
                "exam_type": attempt.exam.mock_test.get_exam_type_display(),
                "overall_band_score": scores["overall_score"],
                "listening_score": scores["listening_score"],
                "reading_score": scores["reading_score"],
                "writing_score": scores["writing_score"],
                "speaking_score": scores["speaking_score"],
                "completed_at": attempt.completed_at,
            }
        )

    serializer = UserDetailSerializer(student)

    return Response(
        {
            "student": serializer.data,
            "results": results_data,
        }
    )


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def student_toggle_active(request, user_id):
    """Toggle student active status"""
    if not check_manager_permission(request.user):
        return permission_denied_response()

    student = get_object_or_404(User, id=user_id, role="STUDENT")

    student.is_active = not student.is_active
    student.save()

    serializer = UserSerializer(student)
    return Response(serializer.data)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def create_student(request):
    """Create a new student"""
    if not check_manager_permission(request.user):
        return permission_denied_response()

    data = request.data.copy()

    # Normalize phone number
    if "phone" in data and data["phone"]:
        data["phone"] = normalize_phone(data["phone"])

    # Clean password - remove any non-alphanumeric characters
    if "password" in data and data["password"]:
        password = data["password"]
        if isinstance(password, (int, float)):
            password = str(int(password))
        else:
            password = str(password)
        # Remove all non-alphanumeric characters (keep only letters and numbers)
        data["password"] = re.sub(r"[^a-zA-Z0-9]", "", password)

    data["role"] = "STUDENT"

    # Generate username if not provided
    if "username" not in data or not data["username"]:
        # Priority: phone > first_name.last_name
        if data.get("phone"):
            base_username = data["phone"]
        elif data.get("first_name") and data.get("last_name"):
            # Generate from first_name and last_name
            first = data["first_name"].lower().strip()
            last = data["last_name"].lower().strip()
            base_username = f"{first}.{last}"
        elif data.get("first_name"):
            base_username = data["first_name"].lower().strip()
        else:
            return Response(
                {"error": "Either phone or first_name is required"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Ensure unique username
        username = base_username
        counter = 1
        while User.objects.filter(username=username).exists():
            username = f"{base_username}{counter}"
            counter += 1
        data["username"] = username

    # Check for duplicate phone
    if data.get("phone") and User.objects.filter(phone=data["phone"]).exists():
        return Response(
            {"error": f"Phone number {data['phone']} already exists"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    serializer = UserSerializer(data=data)
    if serializer.is_valid():
        # Create user with password
        user = User.objects.create_user(
            username=data["username"],
            email=data.get("email", ""),
            password=data.get("password", "defaultpass123"),
            first_name=data.get("first_name", ""),
            last_name=data.get("last_name", ""),
            phone=data.get("phone", ""),
            role="STUDENT",
            is_active=data.get("is_active", True),
        )

        result_serializer = UserSerializer(user)
        return Response(result_serializer.data, status=status.HTTP_201_CREATED)

    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(["PUT"])
@permission_classes([IsAuthenticated])
def update_student(request, user_id):
    """Update student information"""
    if not check_manager_permission(request.user):
        return permission_denied_response()

    student = get_object_or_404(User, id=user_id, role="STUDENT")

    data = request.data.copy()

    # Update fields
    if "first_name" in data:
        student.first_name = data["first_name"]
    if "last_name" in data:
        student.last_name = data["last_name"]
    if "phone" in data:
        normalized_phone = normalize_phone(data["phone"])
        # Check for duplicate phone (excluding current student)
        if User.objects.filter(phone=normalized_phone).exclude(id=student.id).exists():
            return Response(
                {"error": f"Phone number {normalized_phone} already exists"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        student.phone = normalized_phone
    if "is_active" in data:
        student.is_active = data["is_active"]

    student.save()

    serializer = UserSerializer(student)
    return Response(serializer.data)


@api_view(["DELETE"])
@permission_classes([IsAuthenticated])
def delete_student(request, user_id):
    """Delete a student"""
    if not check_manager_permission(request.user):
        return permission_denied_response()

    student = get_object_or_404(User, id=user_id, role="STUDENT")

    student.delete()

    return Response(
        {"message": "Student deleted successfully"}, status=status.HTTP_200_OK
    )


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def bulk_create_students(request):
    """Bulk create students from CSV data"""
    if not check_manager_permission(request.user):
        return permission_denied_response()

    students_data = request.data.get("students", [])

    if not students_data:
        return Response(
            {"error": "No student data provided"}, status=status.HTTP_400_BAD_REQUEST
        )

    created = 0
    failed = 0
    errors = []

    for idx, student_data in enumerate(students_data):
        try:
            # Validate required fields
            if not all(
                k in student_data
                for k in ["first_name", "last_name", "phone", "password"]
            ):
                errors.append(
                    f"Row {idx + 1}: Missing required fields (first_name, last_name, phone, password)"
                )
                failed += 1
                continue

            # Normalize phone number
            phone = normalize_phone(student_data["phone"])
            student_data["phone"] = phone

            # Clean password - remove any non-alphanumeric characters
            password = student_data["password"]
            if isinstance(password, (int, float)):
                password = str(int(password))
            else:
                password = str(password)
            # Remove all non-alphanumeric characters (keep only letters and numbers)
            password = re.sub(r"[^a-zA-Z0-9]", "", password)

            if not password:
                errors.append(f"Row {idx + 1}: Invalid password format")
                failed += 1
                continue

            # Generate unique username from first_name.last_name
            first = student_data["first_name"].lower().strip()
            last = student_data["last_name"].lower().strip()
            base_username = f"{first}.{last}"
            username = base_username
            counter = 1
            while User.objects.filter(username=username).exists():
                username = f"{base_username}{counter}"
                counter += 1

            # Check if phone already exists
            if User.objects.filter(phone=phone).exists():
                errors.append(f"Row {idx + 1}: Phone {phone} already exists")
                failed += 1
                continue

            # Create user
            User.objects.create_user(
                username=username,
                email="",
                password=password,  # Use cleaned password
                first_name=student_data["first_name"],
                last_name=student_data["last_name"],
                phone=phone,
                role="STUDENT",
                is_active=student_data.get("is_active", True),
            )
            created += 1

        except Exception as e:
            errors.append(f"Row {idx + 1}: {str(e)}")
            failed += 1

    return Response(
        {"created": created, "failed": failed, "errors": errors},
        status=(
            status.HTTP_200_OK
            if created > 0 or failed > 0
            else status.HTTP_400_BAD_REQUEST
        ),
    )


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def upload_students_excel(request):
    """Upload Excel file with students data"""
    if not check_manager_permission(request.user):
        return permission_denied_response()

    if "file" not in request.FILES:
        return Response(
            {"error": "No file uploaded"}, status=status.HTTP_400_BAD_REQUEST
        )

    excel_file = request.FILES["file"]

    # Validate file extension
    if not excel_file.name.endswith((".xlsx", ".xls")):
        return Response(
            {
                "error": "Invalid file format. Please upload an Excel file (.xlsx or .xls)"
            },
            status=status.HTTP_400_BAD_REQUEST,
        )

    try:
        # Load workbook
        wb = load_workbook(excel_file, data_only=True)
        ws = wb.active

        created = 0
        failed = 0
        errors = []
        # Skip header row (row 1) and instruction rows (rows 2-3)
        for row_idx, row in enumerate(
            ws.iter_rows(min_row=2, values_only=True), start=2
        ):
            # Skip empty rows
            if not any(row):
                continue

            try:
                first_name, last_name, phone, password = row[:4]
                print(first_name, last_name, phone, password)
                # Validate required fields
                if not all([first_name, last_name, phone, password]):
                    errors.append(f"Row {row_idx}: Missing required fields")
                    failed += 1
                    continue

                # Clean password - convert float to int, then remove any symbols
                # Excel reads numeric passwords as floats (e.g., 12345678.0)
                if isinstance(password, (int, float)):
                    password = str(int(password))  # Convert to int first to remove .0
                else:
                    password = str(password)
                # Remove all non-alphanumeric characters (keep only letters and numbers)
                password = re.sub(r"[^a-zA-Z0-9]", "", password)

                if not password:
                    errors.append(f"Row {row_idx}: Invalid password format")
                    failed += 1
                    continue
                # Normalize phone number
                phone = normalize_phone(str(phone))
                # Generate unique username from first_name.last_name
                first = re.sub(r"[^a-zA-Z0-9]", "", (str(first_name).lower().strip()))
                last = re.sub(r"[^a-zA-Z0-9]", "", (str(last_name).lower().strip()))
                base_username = f"{first}.{last}"
                username = base_username
                counter = 1
                while User.objects.filter(username=username).exists():
                    username = f"{base_username}{counter}"
                    counter += 1

                # Check if phone already exists
                if User.objects.filter(phone=phone).exists():
                    errors.append(f"Row {row_idx}: Phone {phone} already exists")
                    failed += 1
                    continue

                # Create user
                User.objects.create_user(
                    username=username,
                    email="",
                    password=str(password),
                    first_name=str(first_name),
                    last_name=str(last_name),
                    phone=phone,
                    role="STUDENT",
                    is_active=True,
                )
                created += 1

            except Exception as e:
                errors.append(f"Row {row_idx}: {str(e)}")
                failed += 1

        return Response(
            {
                "created": created,
                "failed": failed,
                "errors": errors,
                "message": f"Successfully created {created} students",
            },
            status=(
                status.HTTP_200_OK
                if created > 0 or failed > 0
                else status.HTTP_400_BAD_REQUEST
            ),
        )

    except Exception as e:
        return Response(
            {"error": f"Error processing Excel file: {str(e)}"},
            status=status.HTTP_400_BAD_REQUEST,
        )


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def download_excel_template(request):
    """Download Excel template for bulk student upload"""
    if not check_manager_permission(request.user):
        return permission_denied_response()

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Students Template"

    # Title
    ws["A1"] = "Student Bulk Upload Template"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A1"].fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    ws["A1"].font = Font(size=14, bold=True, color="FFFFFF")
    ws.merge_cells("A1:D1")

    # Instructions
    ws["A2"] = (
        "Instructions: Fill in the student information below. All fields are required. Passwords should contain only letters and numbers (no symbols)."
    )
    ws["A2"].font = Font(italic=True)
    ws.merge_cells("A2:D2")

    # Headers
    headers = ["First Name", "Last Name", "Phone", "Password"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(
            start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
        )
        cell.alignment = Alignment(horizontal="center")

    # Sample data
    sample_data = [
        ["John", "Doe", "998901234567", "password123"],
        ["Jane", "Smith", "998907654321", "secure456"],
    ]

    for row_idx, data in enumerate(sample_data, start=4):
        for col_idx, value in enumerate(data, start=1):
            ws.cell(row=row_idx, column=col_idx).value = value

    # Adjust column widths
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20

    # Prepare response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="students_template.xlsx"'
    wb.save(response)

    return response
