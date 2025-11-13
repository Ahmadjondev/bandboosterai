"""
Manager API - Exam Management Endpoints
Scheduled exams that students can attend with PIN codes
"""

from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.shortcuts import get_object_or_404
from django.db.models import Q, Count, Avg
from django.utils import timezone
from django.core.exceptions import ValidationError
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from ielts.models import Exam, MockExam, Question, UserAnswer, TestHead
from ielts.analysis import calculate_band_score
from ..serializers import ExamSerializer, ExamDetailSerializer
from .utils import (
    check_manager_permission,
    permission_denied_response,
    paginate_queryset,
)


def _calculate_attempt_scores(attempt):
    """
    Calculate scores for an exam attempt by analyzing user answers.
    Returns dict with listening_score, reading_score, writing_score, speaking_score, overall_score
    """
    exam = attempt.exam
    exam_type = exam.mock_test.exam_type
    scores = {
        "listening_score": None,
        "reading_score": None,
        "writing_score": None,
        "speaking_score": None,
        "overall_score": None,
    }

    # Try to get from ExamResult first
    try:
        result = attempt.result
        if result:
            scores["listening_score"] = result.listening_score
            scores["reading_score"] = result.reading_score
            scores["writing_score"] = result.writing_score
            scores["speaking_score"] = result.speaking_score
            scores["overall_score"] = result.overall_score
            return scores
    except:
        pass

    # Calculate scores dynamically if no result exists

    # Listening score
    if exam_type in [
        "LISTENING",
        "LISTENING_READING",
        "LISTENING_READING_WRITING",
        "FULL_TEST",
    ]:
        try:
            listening_questions = Question.objects.filter(
                test_head__listening__in=exam.mock_test.listening_parts.all()
            )

            user_answers = UserAnswer.objects.filter(
                exam_attempt=attempt, question__in=listening_questions
            )

            # Calculate correct answers (considering MCMA weighted scoring)
            correct_count = 0
            total_count = 0

            for q in listening_questions:
                user_answer_obj = user_answers.filter(question=q).first()
                user_answer = (
                    user_answer_obj.answer_text.strip().upper()
                    if user_answer_obj
                    else ""
                )
                correct_answer = (q.get_correct_answer() or "").strip().upper()

                if (
                    q.test_head.question_type
                    == TestHead.QuestionType.MULTIPLE_CHOICE_MULTIPLE_ANSWERS
                ):
                    # MCMA scoring
                    user_set = set(user_answer)
                    correct_set = set(correct_answer)

                    correct_selections = len(user_set & correct_set)
                    incorrect_selections = len(user_set - correct_set)
                    score = max(0, correct_selections - incorrect_selections)

                    correct_count += score
                    total_count += len(correct_set) if correct_set else 1
                else:
                    # Regular scoring
                    if user_answer == correct_answer:
                        correct_count += 1
                    total_count += 1

            if total_count > 0:
                scores["listening_score"] = calculate_band_score(
                    correct_count, total_count, "listening"
                )
        except Exception as e:
            print(f"Error calculating listening score: {e}")

    # Reading score
    if exam_type in [
        "READING",
        "LISTENING_READING",
        "LISTENING_READING_WRITING",
        "FULL_TEST",
    ]:
        try:
            reading_questions = Question.objects.filter(
                test_head__reading__in=exam.mock_test.reading_passages.all()
            )

            user_answers = UserAnswer.objects.filter(
                exam_attempt=attempt, question__in=reading_questions
            )

            # Calculate correct answers (considering MCMA weighted scoring)
            correct_count = 0
            total_count = 0

            for q in reading_questions:
                user_answer_obj = user_answers.filter(question=q).first()
                user_answer = (
                    user_answer_obj.answer_text.strip().upper()
                    if user_answer_obj
                    else ""
                )
                correct_answer = (q.get_correct_answer() or "").strip().upper()

                if (
                    q.test_head.question_type
                    == TestHead.QuestionType.MULTIPLE_CHOICE_MULTIPLE_ANSWERS
                ):
                    # MCMA scoring
                    user_set = set(user_answer)
                    correct_set = set(correct_answer)

                    correct_selections = len(user_set & correct_set)
                    incorrect_selections = len(user_set - correct_set)
                    score = max(0, correct_selections - incorrect_selections)

                    correct_count += score
                    total_count += len(correct_set) if correct_set else 1
                else:
                    # Regular scoring
                    if user_answer == correct_answer:
                        correct_count += 1
                    total_count += 1

            if total_count > 0:
                scores["reading_score"] = calculate_band_score(
                    correct_count, total_count, "reading"
                )
        except Exception as e:
            print(f"Error calculating reading score: {e}")

    # Writing and Speaking scores would need AI evaluation
    # For now, check if there are attempts
    if exam_type in ["WRITING", "LISTENING_READING_WRITING", "FULL_TEST"]:
        from ielts.models import WritingAttempt

        writing_attempts = WritingAttempt.objects.filter(exam_attempt=attempt)
        if writing_attempts.exists():
            writing_scores = [wa.band_score for wa in writing_attempts if wa.band_score]
            if writing_scores:
                scores["writing_score"] = round(
                    sum(writing_scores) / len(writing_scores), 1
                )

    if exam_type in ["SPEAKING", "FULL_TEST"]:
        from ielts.models import SpeakingAttempt

        speaking_attempts = SpeakingAttempt.objects.filter(exam_attempt=attempt)
        if speaking_attempts.exists():
            speaking_scores = [
                sa.band_score for sa in speaking_attempts if sa.band_score
            ]
            if speaking_scores:
                scores["speaking_score"] = round(
                    sum(speaking_scores) / len(speaking_scores), 1
                )

    # Calculate overall score (must be in half-band increments)
    valid_scores = []
    for key in (
        "listening_score",
        "reading_score",
        "writing_score",
        "speaking_score",
    ):
        val = scores.get(key)
        if val is None:
            continue
        try:
            valid_scores.append(float(val))
        except (TypeError, ValueError):
            # Skip values that cannot be converted to float
            continue
    if valid_scores:
        # Average and round to nearest 0.5 (IELTS band score format)
        avg = sum(valid_scores) / len(valid_scores)
        scores["overall_score"] = round(avg * 2) / 2

    return scores


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_exams_list(request):
    """Get list of scheduled exams"""
    if not check_manager_permission(request.user):
        return permission_denied_response()

    search = request.GET.get("search", "")
    status_filter = request.GET.get("status", "")

    exams = (
        Exam.objects.all()
        .select_related("mock_test", "created_by")
        .prefetch_related("enrolled_students")
    )

    if search:
        exams = exams.filter(
            Q(name__icontains=search)
            | Q(mock_test__title__icontains=search)
            | Q(pin_code__icontains=search)
        )

    if status_filter:
        exams = exams.filter(status=status_filter)

    exams = exams.order_by("-start_date")

    paginated = paginate_queryset(exams, request, per_page=15)
    serializer = ExamSerializer(paginated["results"], many=True)

    return Response({"exams": serializer.data, "pagination": paginated["pagination"]})


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_exam_detail(request, exam_id):
    """Get detailed exam information"""
    if not check_manager_permission(request.user):
        return permission_denied_response()

    exam = get_object_or_404(
        Exam.objects.select_related("mock_test", "created_by").prefetch_related(
            "enrolled_students"
        ),
        id=exam_id,
    )

    serializer = ExamDetailSerializer(exam)
    return Response(serializer.data)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def create_exam(request):
    """Create a new scheduled exam"""
    if not check_manager_permission(request.user):
        return permission_denied_response()

    data = request.data.copy()
    data["created_by"] = request.user.id

    mock_test_id = data.get("mock_test")
    if mock_test_id:
        mock_test = get_object_or_404(MockExam, id=mock_test_id)

    serializer = ExamSerializer(data=data)
    if serializer.is_valid():
        try:
            exam = serializer.save(created_by=request.user)
            detail_serializer = ExamDetailSerializer(exam)
            return Response(detail_serializer.data, status=status.HTTP_201_CREATED)
        except ValidationError as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(["PUT", "PATCH"])
@permission_classes([IsAuthenticated])
def update_exam(request, exam_id):
    """Update an existing exam"""
    if not check_manager_permission(request.user):
        return permission_denied_response()

    exam = get_object_or_404(Exam, id=exam_id)

    # # Don't allow updating if exam has started and has enrolled students
    # if exam.enrolled_count > 0 and exam.start_date <= timezone.now():
    #     return Response(
    #         {
    #             "error": "Cannot update exam that has already started with enrolled students"
    #         },
    #         status=status.HTTP_400_BAD_REQUEST,
    #     )

    partial = request.method == "PATCH"
    serializer = ExamSerializer(exam, data=request.data, partial=partial)

    if serializer.is_valid():
        try:
            serializer.save()
            detail_serializer = ExamDetailSerializer(exam)
            return Response(detail_serializer.data)
        except ValidationError as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(["DELETE"])
@permission_classes([IsAuthenticated])
def delete_exam(request, exam_id):
    """Delete an exam"""
    if not check_manager_permission(request.user):
        return permission_denied_response()

    exam = get_object_or_404(Exam, id=exam_id)

    # Don't allow deleting if exam has enrolled students
    if exam.enrolled_count > 0:
        return Response(
            {"error": "Cannot delete exam with enrolled students"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    exam.delete()
    return Response(
        {"message": "Exam deleted successfully"}, status=status.HTTP_204_NO_CONTENT
    )


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def toggle_exam_status(request, exam_id):
    """Toggle exam status (activate/deactivate)"""
    if not check_manager_permission(request.user):
        return permission_denied_response()

    exam = get_object_or_404(Exam, id=exam_id)

    new_status = request.data.get("status")
    if new_status not in dict(Exam.STATUS_CHOICES):
        return Response({"error": "Invalid status"}, status=status.HTTP_400_BAD_REQUEST)

    exam.status = new_status
    exam.save()

    serializer = ExamDetailSerializer(exam)
    return Response(serializer.data)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def remove_student_from_exam(request, exam_id, student_id):
    """Remove a student from an exam"""
    if not check_manager_permission(request.user):
        return permission_denied_response()

    exam = get_object_or_404(Exam, id=exam_id)

    # Don't allow removing if exam has already started
    if exam.start_date <= timezone.now():
        return Response(
            {"error": "Cannot remove student from exam that has already started"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    from django.contrib.auth import get_user_model

    User = get_user_model()
    student = get_object_or_404(User, id=student_id, role="STUDENT")

    if student not in exam.enrolled_students.all():
        return Response(
            {"error": "Student is not enrolled in this exam"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    exam.enrolled_students.remove(student)

    serializer = ExamDetailSerializer(exam)
    return Response(serializer.data)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_exam_statistics(request):
    """Get exam statistics for dashboard"""
    if not check_manager_permission(request.user):
        return permission_denied_response()

    total_exams = Exam.objects.all().count()
    active_exams = Exam.objects.filter(status="ACTIVE").count()
    upcoming_exams = Exam.objects.filter(
        status="SCHEDULED", start_date__gt=timezone.now()
    ).count()
    completed_exams = Exam.objects.filter(status="COMPLETED").count()

    total_enrolled = (
        Exam.objects.all()
        .annotate(count=Count("enrolled_students"))
        .aggregate(total=Count("enrolled_students"))["total"]
        or 0
    )

    return Response(
        {
            "total_exams": total_exams,
            "active_exams": active_exams,
            "upcoming_exams": upcoming_exams,
            "completed_exams": completed_exams,
            "total_enrolled": total_enrolled,
        }
    )


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_exam_results(request, exam_id):
    """Get student results for an exam"""
    if not check_manager_permission(request.user):
        return permission_denied_response()

    # Get exam
    exam = get_object_or_404(Exam, id=exam_id)

    # Get completed attempts for enrolled students
    # Query ExamAttempt instead of ExamResult since results may not be created yet
    from ielts.models import ExamAttempt

    completed_attempts = (
        ExamAttempt.objects.filter(
            exam=exam,
            student__in=exam.enrolled_students.all(),
            status="COMPLETED",
        )
        .select_related("student", "exam")
        .order_by("-completed_at")
    )

    # Prepare results data from completed attempts
    results_data = []
    listening_scores = []
    reading_scores = []
    writing_scores = []
    speaking_scores = []
    total_scores = []

    for attempt in completed_attempts:
        # Calculate scores dynamically from attempt
        scores = _calculate_attempt_scores(attempt)

        # Calculate time spent
        time_spent = None
        if attempt.started_at and attempt.completed_at:
            time_spent = int(
                (attempt.completed_at - attempt.started_at).total_seconds() / 60
            )

        # Get scores
        listening_score = scores["listening_score"]
        reading_score = scores["reading_score"]
        writing_score = scores["writing_score"]
        speaking_score = scores["speaking_score"]
        overall_score = scores["overall_score"]

        # Add to results data
        results_data.append(
            {
                "id": attempt.id,
                "student_name": attempt.student.get_full_name()
                or attempt.student.username,
                "student_email": attempt.student.email,
                "total_score": overall_score,
                "listening_score": listening_score,
                "reading_score": reading_score,
                "writing_score": writing_score,
                "speaking_score": speaking_score,
                "completed_at": attempt.completed_at,
                "time_spent": time_spent,
                "status": "completed",
            }
        )

        # Collect scores for statistics
        if listening_score:
            listening_scores.append(float(listening_score))
        if reading_score:
            reading_scores.append(float(reading_score))
        if writing_score:
            writing_scores.append(float(writing_score))
        if speaking_score:
            speaking_scores.append(float(speaking_score))
        if overall_score:
            total_scores.append(float(overall_score))

    # Get students who haven't completed yet
    completed_user_ids = completed_attempts.values_list("student_id", flat=True)
    pending_students = exam.enrolled_students.exclude(id__in=completed_user_ids)

    pending_data = []
    for student in pending_students:
        pending_data.append(
            {
                "id": None,
                "student_name": student.get_full_name() or student.username,
                "student_email": student.email,
                "total_score": None,
                "listening_score": None,
                "reading_score": None,
                "writing_score": None,
                "speaking_score": None,
                "completed_at": None,
                "time_spent": None,
                "status": "pending",
            }
        )

    # Calculate statistics from collected scores
    avg_total = round(sum(total_scores) / len(total_scores), 2) if total_scores else 0
    avg_listening = (
        round(sum(listening_scores) / len(listening_scores), 2)
        if listening_scores
        else 0
    )
    avg_reading = (
        round(sum(reading_scores) / len(reading_scores), 2) if reading_scores else 0
    )
    avg_writing = (
        round(sum(writing_scores) / len(writing_scores), 2) if writing_scores else 0
    )
    avg_speaking = (
        round(sum(speaking_scores) / len(speaking_scores), 2) if speaking_scores else 0
    )

    return Response(
        {
            "exam": {
                "id": exam.id,
                "name": exam.name,
                "enrolled_count": exam.enrolled_count,
            },
            "completed_results": results_data,
            "pending_students": pending_data,
            "statistics": {
                "total_enrolled": exam.enrolled_count,
                "completed": completed_attempts.count(),
                "pending": pending_students.count(),
                "avg_total_score": avg_total,
                "avg_listening": avg_listening,
                "avg_reading": avg_reading,
                "avg_writing": avg_writing,
                "avg_speaking": avg_speaking,
            },
        }
    )


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def export_exam_results(request, exam_id):
    """Export exam results to Excel file"""
    if not check_manager_permission(request.user):
        return permission_denied_response()

    # Get exam
    exam = get_object_or_404(Exam, id=exam_id)

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Exam Results"

    # Define header style
    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Write exam info
    ws["A1"] = "Exam Name:"
    ws["B1"] = exam.name
    ws["A2"] = "Mock Test:"
    ws["B2"] = exam.mock_test.title
    ws["A3"] = "Export Date:"
    ws["B3"] = timezone.now().strftime("%Y-%m-%d %H:%M")
    ws["A4"] = "Total Enrolled:"
    ws["B4"] = exam.enrolled_count

    # Style exam info
    for row in range(1, 5):
        ws[f"A{row}"].font = Font(bold=True)

    # Headers
    headers = [
        "№",
        "Student Name",
        "Email",
        "Total Score",
        "Listening",
        "Reading",
        "Writing",
        "Speaking",
        "Status",
        "Completed At",
        "Time Spent (min)",
    ]

    header_row = 6
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Get results from completed attempts
    # Query ExamAttempt instead of ExamResult since results may not be created yet
    from ielts.models import ExamAttempt

    completed_attempts = (
        ExamAttempt.objects.filter(
            exam=exam,
            student__in=exam.enrolled_students.all(),
            status="COMPLETED",
        )
        .select_related("student")
        .order_by("-completed_at", "student__first_name")
    )

    # Write completed results
    row = header_row + 1
    for idx, attempt in enumerate(completed_attempts, start=1):
        # Calculate scores dynamically from attempt
        scores = _calculate_attempt_scores(attempt)

        # Calculate time spent
        time_spent = ""
        if attempt.started_at and attempt.completed_at:
            time_spent = int(
                (attempt.completed_at - attempt.started_at).total_seconds() / 60
            )

        # Get scores
        listening_score = scores["listening_score"]
        reading_score = scores["reading_score"]
        writing_score = scores["writing_score"]
        speaking_score = scores["speaking_score"]
        overall_score = scores["overall_score"]

        ws.cell(row=row, column=1, value=idx)
        ws.cell(
            row=row,
            column=2,
            value=attempt.student.get_full_name() or attempt.student.username,
        )
        ws.cell(row=row, column=3, value=attempt.student.email)
        ws.cell(
            row=row,
            column=4,
            value=float(overall_score) if overall_score else "-",
        )
        ws.cell(
            row=row,
            column=5,
            value=float(listening_score) if listening_score else "-",
        )
        ws.cell(
            row=row,
            column=6,
            value=float(reading_score) if reading_score else "-",
        )
        ws.cell(
            row=row,
            column=7,
            value=float(writing_score) if writing_score else "-",
        )
        ws.cell(
            row=row,
            column=8,
            value=float(speaking_score) if speaking_score else "-",
        )
        ws.cell(row=row, column=9, value="Completed")
        ws.cell(
            row=row,
            column=10,
            value=(
                attempt.completed_at.strftime("%Y-%m-%d %H:%M")
                if attempt.completed_at
                else ""
            ),
        )
        ws.cell(row=row, column=11, value=time_spent)
        row += 1

    # Write pending students
    completed_user_ids = completed_attempts.values_list("student_id", flat=True)
    pending_students = exam.enrolled_students.exclude(
        id__in=completed_user_ids
    ).order_by("first_name")

    for student in pending_students:
        ws.cell(row=row, column=1, value=row - header_row)
        ws.cell(row=row, column=2, value=student.get_full_name() or student.username)
        ws.cell(row=row, column=3, value=student.email)
        ws.cell(row=row, column=4, value="-")
        ws.cell(row=row, column=5, value="-")
        ws.cell(row=row, column=6, value="-")
        ws.cell(row=row, column=7, value="-")
        ws.cell(row=row, column=8, value="-")
        ws.cell(row=row, column=9, value="Pending")
        ws.cell(row=row, column=10, value="-")
        ws.cell(row=row, column=11, value="-")
        row += 1

    # Add statistics
    if completed_attempts.exists():
        stats_row = row + 2
        ws.cell(row=stats_row, column=1, value="STATISTICS")
        ws.cell(row=stats_row, column=1).font = Font(bold=True, size=12)

        stats_row += 1
        ws.cell(row=stats_row, column=1, value="Average Scores:")
        ws.cell(row=stats_row, column=1).font = Font(bold=True)

        # Calculate averages from attempts
        listening_scores = []
        reading_scores = []
        writing_scores = []
        speaking_scores = []
        total_scores = []

        for attempt in completed_attempts:
            scores = _calculate_attempt_scores(attempt)

            if scores["listening_score"]:
                listening_scores.append(float(scores["listening_score"]))
            if scores["reading_score"]:
                reading_scores.append(float(scores["reading_score"]))
            if scores["writing_score"]:
                writing_scores.append(float(scores["writing_score"]))
            if scores["speaking_score"]:
                speaking_scores.append(float(scores["speaking_score"]))
            if scores["overall_score"]:
                total_scores.append(float(scores["overall_score"]))

        avg_total = (
            round(sum(total_scores) / len(total_scores), 2) if total_scores else 0
        )
        avg_listening = (
            round(sum(listening_scores) / len(listening_scores), 2)
            if listening_scores
            else 0
        )
        avg_reading = (
            round(sum(reading_scores) / len(reading_scores), 2) if reading_scores else 0
        )
        avg_writing = (
            round(sum(writing_scores) / len(writing_scores), 2) if writing_scores else 0
        )
        avg_speaking = (
            round(sum(speaking_scores) / len(speaking_scores), 2)
            if speaking_scores
            else 0
        )

        stats_row += 1
        ws.cell(row=stats_row, column=1, value="Total:")
        ws.cell(row=stats_row, column=2, value=avg_total)

        stats_row += 1
        ws.cell(row=stats_row, column=1, value="Listening:")
        ws.cell(row=stats_row, column=2, value=avg_listening)

        stats_row += 1
        ws.cell(row=stats_row, column=1, value="Reading:")
        ws.cell(row=stats_row, column=2, value=avg_reading)

        stats_row += 1
        ws.cell(row=stats_row, column=1, value="Writing:")
        ws.cell(row=stats_row, column=2, value=avg_writing)

        stats_row += 1
        ws.cell(row=stats_row, column=1, value="Speaking:")
        ws.cell(row=stats_row, column=2, value=avg_speaking)

    # Adjust column widths
    column_widths = [5, 25, 30, 12, 12, 12, 12, 12, 15, 20, 18]
    for col, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Create response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"exam_{exam.id}_{exam.name.replace(' ', '_')}_results.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_student_result_detail(request, attempt_id):
    """Get detailed student result for a specific attempt - Manager Access"""
    if not check_manager_permission(request.user):
        return permission_denied_response()

    # Import here to avoid circular imports
    from ielts.models import ExamAttempt
    from ielts.api_views import (
        _get_listening_results,
        _get_reading_results,
        _get_writing_results,
        _get_speaking_results,
    )
    from ielts.analysis import identify_strengths_and_weaknesses

    try:
        attempt = ExamAttempt.objects.select_related(
            "student", "exam", "exam__mock_test"
        ).get(id=attempt_id)
    except ExamAttempt.DoesNotExist:
        return Response(
            {"error": "Exam attempt not found or access denied"},
            status=status.HTTP_404_NOT_FOUND,
        )

    if attempt.status != "COMPLETED":
        return Response(
            {"error": "Test has not been completed yet."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    exam = attempt.exam
    exam_type = exam.mock_test.exam_type

    results = {
        "attempt_id": attempt.id,
        "exam_title": exam.mock_test.title,
        "exam_type": exam_type,
        "completed_at": attempt.completed_at,
        "overall_score": _calculate_attempt_scores(attempt).get("overall_score"),
        "duration_minutes": attempt.get_duration_minutes(),
        "student_name": attempt.student.get_full_name() or attempt.student.username,
        "student_email": attempt.student.email,
        "sections": {},
    }

    analysis_results = []

    # Listening Section
    if exam_type in [
        "LISTENING",
        "LISTENING_READING",
        "LISTENING_READING_WRITING",
        "FULL_TEST",
    ]:
        listening_data = _get_listening_results(attempt)
        results["sections"]["listening"] = listening_data
        analysis_results.append(
            {
                "section_name": "Listening",
                "accuracy_by_type": listening_data.get("accuracy_by_type", {}),
                "accuracy_by_part": listening_data.get("accuracy_by_part", {}),
            }
        )

    # Reading Section
    if exam_type in [
        "READING",
        "LISTENING_READING",
        "LISTENING_READING_WRITING",
        "FULL_TEST",
    ]:
        reading_data = _get_reading_results(attempt)
        results["sections"]["reading"] = reading_data
        analysis_results.append(
            {
                "section_name": "Reading",
                "accuracy_by_type": reading_data.get("accuracy_by_type", {}),
            }
        )

    # Writing Section
    if exam_type in ["WRITING", "LISTENING_READING_WRITING", "FULL_TEST"]:
        writing_data = _get_writing_results(attempt)
        results["sections"]["writing"] = writing_data

    # Speaking Section
    if exam_type in ["SPEAKING", "FULL_TEST"]:
        speaking_data = _get_speaking_results(attempt)
        results["sections"]["speaking"] = speaking_data

    # Overall Analysis
    if analysis_results:
        insights = identify_strengths_and_weaknesses(analysis_results)
        results["insights"] = insights

    return Response(results)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def evaluate_writing_submission(request, attempt_id, task_id):
    """Manager evaluates a student's writing submission with feedback and scores"""
    if not check_manager_permission(request.user):
        return permission_denied_response()

    # Import models
    from ielts.models import ExamAttempt, WritingAttempt

    try:
        attempt = ExamAttempt.objects.get(id=attempt_id)
    except ExamAttempt.DoesNotExist:
        return Response(
            {"error": "Exam attempt not found or access denied"},
            status=status.HTTP_404_NOT_FOUND,
        )

    # Get writing attempt
    try:
        writing_attempt = WritingAttempt.objects.get(id=task_id, exam_attempt=attempt)
    except WritingAttempt.DoesNotExist:
        return Response(
            {"error": "Writing submission not found"},
            status=status.HTTP_404_NOT_FOUND,
        )

    # Extract evaluation data from request
    data = request.data

    # Validate required fields
    required_criteria = [
        "task_response_or_achievement",
        "coherence_and_cohesion",
        "lexical_resource",
        "grammatical_range_and_accuracy",
    ]

    for criterion in required_criteria:
        if criterion not in data:
            return Response(
                {"error": f"Missing required field: {criterion}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Validate score range (0-9, increments of 0.5)
        try:
            score = float(data[criterion])
            if score < 0 or score > 9:
                return Response(
                    {
                        "error": f"{criterion} must be between 0 and 9",
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )
            # Check if it's in 0.5 increments
            if (score * 2) % 1 != 0:
                return Response(
                    {
                        "error": f"{criterion} must be in 0.5 increments (e.g., 6.0, 6.5, 7.0)",
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )
        except (ValueError, TypeError):
            return Response(
                {"error": f"Invalid score for {criterion}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

    # Update writing attempt with scores
    writing_attempt.task_response_or_achievement = data["task_response_or_achievement"]
    writing_attempt.coherence_and_cohesion = data["coherence_and_cohesion"]
    writing_attempt.lexical_resource = data["lexical_resource"]
    writing_attempt.grammatical_range_and_accuracy = data[
        "grammatical_range_and_accuracy"
    ]

    # Build feedback JSON
    feedback = {
        "task_response_or_achievement": data.get(
            "feedback_task_response_or_achievement", ""
        ),
        "coherence_and_cohesion": data.get("feedback_coherence_and_cohesion", ""),
        "lexical_resource": data.get("feedback_lexical_resource", ""),
        "grammatical_range_and_accuracy": data.get(
            "feedback_grammatical_range_and_accuracy", ""
        ),
        "overall": data.get("feedback_overall", []),
    }

    writing_attempt.feedback = feedback
    writing_attempt.evaluation_status = WritingAttempt.EvaluationStatus.COMPLETED
    writing_attempt.evaluated_at = timezone.now()

    # The save() method will auto-calculate band_score from the 4 criteria
    writing_attempt.save()

    # Refresh from database to get the calculated band_score
    writing_attempt.refresh_from_db()

    # Update ExamAttempt with writing scores
    from ielts.models import WritingAttempt as WA
    from decimal import Decimal

    # Calculate average writing score from all completed writing tasks
    completed_writing_attempts = WA.objects.filter(
        exam_attempt=attempt, evaluation_status=WA.EvaluationStatus.COMPLETED
    ).exclude(band_score__isnull=True)

    if completed_writing_attempts.exists():
        # Sum Decimal values directly without converting to float
        total_score = sum(wa.band_score for wa in completed_writing_attempts)
        avg_writing_score = total_score / completed_writing_attempts.count()
        # Round to nearest 0.5
        attempt.writing_score = round(avg_writing_score * 2) / 2

        # Recalculate overall score
        attempt.overall_score = attempt.calculate_overall_score()
        attempt.save()

    return Response(
        {
            "message": "Writing evaluation saved successfully",
            "band_score": (
                float(writing_attempt.band_score) if writing_attempt.band_score else 0.0
            ),
            "task_name": writing_attempt.task.get_task_type_display(),
            "writing_score": (
                float(attempt.writing_score) if attempt.writing_score else None
            ),
            "overall_score": (
                float(attempt.overall_score) if attempt.overall_score else None
            ),
        },
        status=status.HTTP_200_OK,
    )
